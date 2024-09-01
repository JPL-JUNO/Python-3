"""
@File         : updateProduce.py
@Author(s)    : Stephen CUI
@LastEditor(s): Stephen CUI
@CreatedTime  : 2024-08-31 17:22:01
@Email        : cuixuanstephen@gmail.com
@Description  : Corrects costs in produce sales spreadsheet.
"""

import openpyxl

wb = openpyxl.load_workbook("data/produceSales.xlsx")
sheet = wb["Sheet"]

PRICE_UPDATES = {
    "Garlic": 3.07,
    "Celery": 1.19,
    "Lemon": 1.27,
}


for row in range(2, sheet.max_row):
    produce_name = sheet.cell(row=row, column=1).value
    if produce_name in PRICE_UPDATES:
        sheet.cell(row=row, column=2).value = PRICE_UPDATES[produce_name]

wb.save("processed/updatedProduceSales.xlsx")
